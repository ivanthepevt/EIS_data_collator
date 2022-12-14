from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import filedialog as fd

def getnewdir(olddir):
    splitdir = olddir.split("/")
    oldfilename = splitdir[-1]
    newfilename = ""
    splitdir.pop()
    for text in splitdir:
        newfilename = newfilename + text + "/"
    newfilename = newfilename + "colatted_" + oldfilename
    return newfilename

def divideListBaseOnName(alist):
    global indicator
    finalist = []
    prev_name = ""
    sublist = ["null"]
    for elem in alist:
        name = elem.split(" ")[0]
        if(name != prev_name):
            prev_name = name
            if(len(sublist) != 0):
                finalist.append(sublist)
                sublist = []
                sublist.append(indicator + name)
                sublist.append(elem)
            else:
                sublist.append(elem)
        else:
            sublist.append(elem)
    if (len(sublist) != 0):
        finalist.append(sublist)
    finalist.pop(0)
    return finalist

indicator = "Collated "
filename = fd.askopenfilename()
print("File chosen: " + filename)
wb = load_workbook(filename, data_only=True)
wb_new = Workbook()

sheetlist = wb.sheetnames
sheetlist.sort()
groupedSheetlist = divideListBaseOnName(sheetlist)
#print(groupedSheetlist)

for group in groupedSheetlist:
    table = []
    groupName = group.pop(0)
    groupedSheet = wb_new.create_sheet(groupName)
    for sheet in group:
        if(indicator+sheet != groupName):
            curentSheet = wb[sheet]
            for col in "ACDEFLMZZ":
                colum = []
                for val in curentSheet[col]:
                    colum.append(val.value)
                #print(colum)
                table.append(colum)

    transposeTable = map(list, zip(*table))
    for row in transposeTable:
        groupedSheet.append(row)
    #groupedSheet.append(headers)

wb_new.save(getnewdir(filename))